#!/usr/bin/python3

import openpyxl as opx
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

#TODO: draw central cells
#TODO: try to print sheet on printer

#

#defined mm var for width of line
wmm5 = 2.016
wmm10 = wmm5 * 2
wmm3 = wmm10 * 0.3
wmm7 = wmm10 * 0.7
wmm8 = wmm10 * 0.8
wmm14 = wmm7 * 2
wmm15 = wmm10 * 1.5
wmm20 = wmm10 * 2
wmm25 = wmm8 * 3
wmm35 = wmm8 * 4

#row height
hmm10 = 28.4
hmm5  = hmm10 / 2
hmm15 = hmm10 * 1.5
hmm25 = hmm10 * 2.5
hmm8 = hmm10 * 0.8

begin_row = 2
end_row = 40
begin_column = 2
end_column = 21
central_table_begin = 3
central_table_end = 31
last_row = 40

cellWidth = {'B':wmm5,'C':wmm7,'D':wmm7,'E':wmm5,'F':wmm5,'G':wmm3,'H':wmm20,'I':wmm15,'J':wmm10,'K':wmm10 * 2.7,
             'L':wmm10 * 3.7,'M':wmm5,'N':wmm5,'O':wmm5,'P':wmm5,'Q':wmm3,'R':wmm10,'S':wmm10*0.2,'T':wmm10,'U':wmm10}

#border style
bold_border_style = Side(border_style="medium") #bold border lines
common_border_style = Side(border_style="hair") #bold border lines

out_border = Border(left=bold_border_style,
                     right=bold_border_style,
                     top=bold_border_style,
                     bottom=bold_border_style)

in_border = Border(left=bold_border_style,
                     right=bold_border_style,
                     top=common_border_style,
                     bottom=common_border_style)


# form cells for GOST table
def setCellWidth(worksheet):
    for cell in cellWidth:
        worksheet.column_dimensions[cell].width = cellWidth[cell]

def setCellHieght(worksheet):
    worksheet.row_dimensions[2].height = hmm15
    for row in range(central_table_begin, last_row):
        if(row < 31):
            worksheet.row_dimensions[row].height = hmm8
        else:
            worksheet.row_dimensions[row].height = hmm5

#set GOST Type A font in cells int rectangle
def setGostFont(worksheet):
    for row in range(2,end_row):
        for column in range(2,end_column+1):
            worksheet.cell(row,column).font = "GOST Type AU"


def left_cells_of_frame(worksheet,column,begin, height,text):
    print(column + begin)
    lst = int(begin) + height
    print(column + str(lst))
    worksheet.merge_cells(column + begin + ':' + column + str(lst))
    top_left_cell = worksheet[column + begin]
    bot_left_cell = worksheet[column + str(lst)]
    # need set this field for top and bot cell,
    # don't know why not work with top cell like in write in documentations
    top_left_cell.border = out_border
    bot_left_cell.border = out_border
    top_left_cell.alignment =Alignment(vertical="center",text_rotation=90) # rotation text to 90 angles
    top_left_cell.value = text

#draw left stamp cells
def draw_left_stamp(worksheet):
    bc_column_height = {'2': 6, '9': 7, '20': 3, '24': 2, '27': 2, '30': 4, '35': 4}  # № of beginin row
    bc_column_text = {'2': "Перв.примен.", '9': "Справ. №", '20': "Подп. и дата", '24': "Инв. № дубл."
        , '27': "Взам. Инв. №", '30': "Подп. и дата", '35': "Инв. № подл."}  # № of beginin row
    # and needing height of result cell
    for row in bc_column_height:
        left_cells_of_frame(worksheet, 'B', row, bc_column_height[row], bc_column_text[row])
        left_cells_of_frame(worksheet, 'C', row, bc_column_height[row], "")

#header is merges from D to U
def draw_header(worksheet):
    #begin:last column to merge
    active_row = 2
    cells = {'D':'G','H':'L','M':'N','O':'U'}
    text = {'D':'Поз.\nОбознач.','H':'Наименование','M':'Кол.','O':'Примечание'}
    for cell in cells:
        begin_cell = cell + str(active_row)
        last_cell = cells[cell] + str(active_row)
        worksheet.merge_cells(begin_cell + ':' + last_cell)
        top_left_cell = worksheet[begin_cell]
        bot_left_cell = worksheet[last_cell]
        top_left_cell.border = out_border
        bot_left_cell.border = out_border
        top_left_cell.value = text[cell]
        if cell == 'M':
            top_left_cell.alignment = Alignment(vertical="center", horizontal="center", text_rotation=90)  # rotation text to 90 angles and text alignment
        else:
            top_left_cell.alignment = Alignment(vertical="center",horizontal="center")  # text alignment




def draw_central_cell(worksheet):
    cells = {'D':'G','H':'L','M':'N','O':'U'}
    for row in range(central_table_begin, central_table_end):
        for cell in cells:
            begin_cell = cell + str(row)
            last_cell = cells[cell] + str(row)
            worksheet.merge_cells(begin_cell + ':' + last_cell)
            top_left_cell = worksheet[begin_cell]
            bot_left_cell = worksheet[last_cell]
            top_left_cell.border = in_border
            bot_left_cell.border = in_border
            top_left_cell.alignment = Alignment(vertical="center", horizontal="center")  # text alignment

workbook_path = "test.xlsx"
workbook = opx.load_workbook(workbook_path)
worksheet = workbook.active

worksheet = workbook.worksheets[0]
'''
worksheet.cell(row=2, column=2).border = bold_border
worksheet.merge_cells('A2:D2')
worksheet.merge_cells('C2:C8')
worksheet.merge_cells('B2:B8')
'''
#set form
setCellWidth(worksheet)
setCellHieght(worksheet)
setGostFont(worksheet)
#mergens
#left cells of page

draw_left_stamp(worksheet)
draw_central_cell(worksheet)
draw_header(worksheet)
#close file with save
workbook.save(workbook_path)
